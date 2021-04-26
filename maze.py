import sys
import os
import copy
import xlsxwriter
from PyQt5.QtWidgets import QDialog, QApplication
from PyQt5.Qt import Qt
from maze_graf import *
from pathlib import Path

import os.path
from stopwatch import Stopwatch
from datetime import datetime

import openpyxl

col=0

class MyInterface(QDialog): 
    # tutti cio che riguarda 'Qdialog' (metodi e valori), valogono anche per MyApp
    esito={} # variabili che salvano i dati durante la sessione. Quando premo fine sessione i dati vengono copiato su un nuovo dizionario. Questi evngono azzersti
    lista=[] # variabili che salvano i dati durante la sessione. Quando premo fine sessione i dati vengono copiato su un nuovo dizionario.Questi evngono azzersti
    ripetizione=''
    now = datetime.now()
    save_path=str(os.getcwd())   #ottengo la working directory del programma
    completeName= 'analisi.txt'  #nome del file dove esporto.

    # Creiamo il cronometro

    stopwatchA = Stopwatch()
    stopwatchB = Stopwatch()
    stopwatchC = Stopwatch()
    stopwatchA.reset()
    stopwatchB.reset()
    stopwatchC.reset()



    f_input = '' #Variabile che contiene il numero di sessione dell'animale
    # esito={} # variabili che salvano i dati durante la sessione. Quando premo fine sessione i dati vengono copiato su un nuovo dizionario. Questi evngono azzersti
    # lista=[] # variabili che salvano i dati durante la sessione. Quando premo fine sessione i dati vengono copiato su un nuovo dizionario. Questi evngono azzersti


    
    def __init__(self):
        super().__init__() # indica quali attiburi da ereditare dall'oggetto passato di classe QDialog
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)
        '''fino a qua è sempre uguale'''



        #BOTTONE FINE SESSIONE
        self.ui.FineSessione.clicked.connect(self.fine_sessione) #collego la funzione 'fine sessione' al bottone fine sessione
        # self.f_input=self.ui.lineEdit_2.text()      # prendo il numero della sessione
        self.lista=[]
        self.esito={}
        self.sessione={}
        self.esportazione={}


        #BOTTONE ESPORTA
        self.ui.Esporta.clicked.connect(self.esporta)

        #BOTTONE INIZIA
        self.ui.inizia.clicked.connect(self.inizia)


    #FUNZIONI PER LA SCELTA DEL BRACCIO
    
    def centro(self,*args):
        self.ripetizione=''
        self.stopwatchA.stop() # Stop it again
        self.stopwatchB.stop() # Stop it again
        self.stopwatchC.stop() # Stop it again
        mini_lista={}
        if self.stopwatchA.duration != 0.0:
            mini_lista['A']=round(self.stopwatchA.duration,1)
        if self.stopwatchB.duration != 0.0:
            mini_lista['B']=round(self.stopwatchB.duration,1)
        if self.stopwatchC.duration != 0.0:
            mini_lista['C']=round(self.stopwatchC.duration,1)     
        self.stopwatchA.reset()
        self.stopwatchB.reset()
        self.stopwatchC.reset()
        if mini_lista !={}:
            self.lista.append(mini_lista)
        # chiave=self.ui.lineEdit_2.text()  # prendo il numero animale
        self.ui.label_stampe.setText('CENTRO')
        #print('Centro')
    
    
    def braccioA(self,*args):
        if self.ripetizione=='A':
            return
        else:
            self.ripetizione='A'
            self.stopwatchA.stop()  # Stop it again
            self.stopwatchB.stop()  # Stop it again
            self.stopwatchC.stop()  # Stop it again
            mini_lista={}
            if self.stopwatchA.duration != 0.0:
                mini_lista['A']=round(self.stopwatchA.duration,1)
            if self.stopwatchB.duration != 0.0:
                mini_lista['B']=round(self.stopwatchB.duration,1)
            if self.stopwatchC.duration != 0.0:
                mini_lista['C']=round(self.stopwatchC.duration,1)
            self.stopwatchA.reset()
            self.stopwatchB.reset()
            self.stopwatchC.reset()
            if mini_lista !={}:
                self.lista.append(mini_lista)
            self.stopwatchA.start() # Start it again
            
            chiave=self.ui.lineEdit_2.text()         # prendo il numero animale
            self.ui.label_stampe.setText('BRACCIO A')
            self.esito['Animale_nr:_'+str(chiave)]=self.lista
            #print('Braccio A')
    
    def braccioB(self,*args):
        if self.ripetizione=='B':
            return
        else:
            self.ripetizione='B'
            self.stopwatchA.stop()  # Stop it again
            self.stopwatchB.stop()  # Stop it again
            self.stopwatchC.stop()  # Stop it again
            mini_lista={}
            if self.stopwatchA.duration != 0.0:
                mini_lista['A']=round(self.stopwatchA.duration,1)
            if self.stopwatchB.duration != 0.0:
                mini_lista['B']=round(self.stopwatchB.duration,1)
            if self.stopwatchC.duration != 0.0:
                mini_lista['C']=round(self.stopwatchC.duration,1)
            self.stopwatchA.reset()
            self.stopwatchB.reset()
            self.stopwatchC.reset()
            if mini_lista !={}:
                self.lista+=[mini_lista]
            self.stopwatchB.start() # Start it again
            chiave=self.ui.lineEdit_2.text()         # prendo il numero animale
            self.ui.label_stampe.setText('BRACCIO B')
            self.esito['Animale_nr:_'+str(chiave)]=self.lista
            #print('Braccio B')
        
    def braccioC(self,*args):
        if self.ripetizione=='C':
            return
        else:
            self.ripetizione='C'
        self.stopwatchA.stop()  # Stop it again
        self.stopwatchB.stop()  # Stop it again
        self.stopwatchC.stop()  # Stop it again
        mini_lista={}
        if self.stopwatchA.duration != 0.0:
            mini_lista['A']=round(self.stopwatchA.duration,1)
        if self.stopwatchB.duration != 0.0:
            mini_lista['B']=round(self.stopwatchB.duration,1)
        if self.stopwatchC.duration != 0.0:
            mini_lista['C']=round(self.stopwatchC.duration,1)
        self.stopwatchA.reset()
        self.stopwatchB.reset()
        self.stopwatchC.reset()
        if mini_lista !={}:
            self.lista+=[mini_lista]
        self.stopwatchC.start() # Start it again
        chiave=self.ui.lineEdit_2.text()         # prendo il numero animale
        self.ui.label_stampe.setText('BRACCIO C')
        self.esito['Animale_nr:_'+str(chiave)]=self.lista
        #print('Braccio C')


    def fsociety(self, lista_da_testo):
        lista_da_testo=lista_da_testo[:-1]
        #print(lista_da_testo)
        os.remove(os.path.join(str(cwd),'analisi_maze.txt'))
        #print('File eliminato')
        return lista_da_testo

    def inizia(self):
        #print('inizio')
        #ammazza(esito,lista)
        cwd = os.getcwd()
        #os.remove(os.path.join(str(cwd),'analisi_maze.txt'))
        #print('ciao')
        tutto=[]
        try:
            with open (cwd+'/analisi_maze.txt', 'r') as f:
                tutto = f.readlines()
                #corretto=fsociety(tutto)
                #print('dentro')
                corretto= tutto[:-1]
                with open (cwd+'/analisi_maze.txt', 'w') as f:    
                    for e in corretto:
                        f.write(e + "\n")
                return
        except:
            pass



    def ammazza(self, esito,lista):
        '''Azzera tutti i valori per far partire la sessione di un nuovo topo'''
        self.esito.clear()
        self.lista.clear()


    #def inizia(self):
        #self.ammazza(self.sessione,self.lista)
        #return

    
    def fine_sessione(self,col):
        '''Termina l'analisi del video aggiungendo FINE come ultimo valore rilevato.
        Annulla anche una sessione se ho fatto un errore. Dopo aver premuto riaprto
        da una nuova sessione.'''
        self.stopwatchA.stop() # Stop it again
        self.stopwatchB.stop() # Stop it again
        self.stopwatchC.stop() # Stop it again
        mini_lista={}
        if self.stopwatchA.duration != 0.0:
            mini_lista['A']=round(self.stopwatchA.duration,1)
        if self.stopwatchB.duration != 0.0:
            mini_lista['B']=round(self.stopwatchB.duration,1)
        if self.stopwatchC.duration != 0.0:
            mini_lista['C']=round(self.stopwatchC.duration,1)     
        self.stopwatchA.reset()
        self.stopwatchB.reset()
        self.stopwatchC.reset()

        if mini_lista !={}:
            self.lista.append(mini_lista)
            stringa=''
        for k,v in self.esito.items():
            stringa+=k
            stringa+=' '
            stringa+=str(v)  # del  tempo elimino la 's'
            stringa+='\n'
        nome_file='analisi_maze.txt'
        save_path=str(os.getcwd())
        completeName= 'analisi_maze.txt'
        scrivere=''
        simboli=["[", "]", "'", "s", ","]
        try:
            for carattere in stringa:
                if carattere in simboli:
                    pass
                else:
                    scrivere+=carattere
            stringa=scrivere
            scrivere=''
            for carattere in stringa:
                if carattere == '.':
                    scrivere += ","
                else:
                    scrivere+=carattere
            stringa=scrivere
            c=0
            for file in os.listdir(save_path):
                if str(file) == nome_file:
                    c+=1
            if c==1:
                with open(os.path.join(save_path,completeName), "a", encoding="utf-8") as f:
                    f.write(stringa)
            else:
                with open(os.path.join(save_path,completeName), "w", encoding="utf-8") as f:
                    f.write(stringa)
            self.ammazza(self.sessione,self.lista)
        except:
            pass


        self.ui.label_stampe.setText('SESSIONE CONCLUSA')
        #print('Sessione conclusa')
        # chiave=self.ui.lineEdit_2.text()
        self.ripetizione=''
        return


    def autoIncrement(self): 
        global col
        pStart = 1
        pInterval = 1
        if (col == 0):
            col = pStart
        else:
            col += pInterval
        return col  

    
    def interruttore(self):
        there = 0
        cwd = os.getcwd()
        for file in os.listdir(cwd):
            if str(file) == 'Analisi_maze.xlsx':
                there=1
            else:
                pass
        return there

    def esporta(self):
        '''Conclude la sessione di un singolo topo e trascrive i risultati in un file txt
        poi li passa in un excel'''
        # print(sessione) #ripendo la sessione creata
        global col
        cwd = os.getcwd()
        workbook = xlsxwriter.Workbook('Analisi_maze.xlsx')
        worksheet = workbook.add_worksheet()
        self.ui.label_stampe.setText('Dati Esportati')
        '''Prende il file di testo delle sessioni e le mette in un excel'''
        with open('analisi_maze.txt', 'r') as f:
            tutto=f.readlines()
        row = 0
        there=self.interruttore()
        if there == 0:
            #print('NON esiste Excel')
            filepath=str(cwd+'/Analisi_maze.xlsx')
            wb = openpyxl.Workbook()
            wb.save(filepath)
            for riga in tutto:
                sep=riga.split(' ')
                for i in range(len(sep)):
                    sep[i]=sep[i].replace('{','')
                    sep[i]=sep[i].replace('}','')
                    sep[i]=sep[i].replace('\n','')
                    worksheet.write(row+i, col, sep[i])
            workbook.close()
            self.autoIncrement()
        else:
            #print('ESISTE Excel')
            #print(col)
            rosetta={1:'A', 2:'B', 3:'C', 4:'D', 5:'E', 6:'F', 7:'G', 8:'H',
            9:'I', 10: 'J', 11:'K', 12:'L', 13: 'M', 14: 'N', 15:'O', 16: 'P',
            17:'Q', 18:'R', 19:'S', 20:'T', 21:'U', 22:'V', 23:'W',24:'X', 25:'Y', 26:'Z',
            27:'AA', 28:'AB', 29:'AC', 30:'AD', 31:'AE', 32:'AF', 33:'AG', 34:'AH',
            35:'AI', 36: 'AJ', 37:'AK', 38:'AL', 39: 'AM', 40: 'AN', 41:'AO', 42: 'AP',
            43:'AQ', 44:'AR', 45:'AS', 46:'AT', 47:'AU', 48:'AV', 49:'AW', 50:'AX', 51:'AY',
            52:'AZ'}
            xfile = openpyxl.load_workbook(cwd+'/Analisi_maze.xlsx')
            sheet = xfile['Sheet1']
            for riga in tutto:
                sep=riga.split(' ')
                for i in range(len(sep)):
                    sep[i]=sep[i].replace('{','')
                    sep[i]=sep[i].replace('}','')
                    sep[i]=sep[i].replace('\n','')
                    sheet[str(rosetta[col+1])+str(i+1)] = sep[i]
            xfile.save('Analisi_maze.xlsx')
            self.autoIncrement()

        return



    # def pulisci(self, stringa):
    #     '''Prende la stringa da scrivere nel file di testo ed elimina i caratter che non servono'''
    #     scrivere=''
    #     simboli=["[", "]", "'", "s", ","]
    #     for carattere in stringa:
    #         if carattere in simboli:
    #             scrivere+=' '
    #         else:
    #             scrivere+= carattere
    #     return scrivere

    def keyPressEvent(self,event):
        '''Questa è la magica funzione che mi permette di prendere gli input da tastiera'''
        if event.key() == Qt.Key_J:
            self.braccioA()
            
        if event.key() == Qt.Key_K:
            self.braccioB()
            
        if event.key() == Qt.Key_L:
            self.braccioC()
            
        if event.key() == Qt.Key_S:
            self.centro()

    def test_method(self):
        #print('Premuto spazio')



        self.show() ##NON ELIMINARE



if __name__=="__main__":
    app = QApplication(sys.argv)
    demo = MyInterface()
    demo.show()
    sys.exit(app.exec_())




workbook = xlsxwriter.Workbook('Expe.xlsx')
worksheet = workbook.add_worksheet()




